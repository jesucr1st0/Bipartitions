# tests/limpiar_excel.py
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
import openpyxl

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
EXCEL_PATH   = os.path.join(PROJECT_ROOT, "data", "DatosPruebas2026_1.xlsx")

FILA_INICIO = 6
COLUMNAS_GEOMETRIC = [7, 8, 9, 13, 14, 15, 19, 20, 21, 25, 26, 27]

wb = openpyxl.load_workbook(EXCEL_PATH)

for hoja in ["10A-Elementos", "15B-Elementos"]:
    if hoja not in wb.sheetnames:
        continue
    ws = wb[hoja]
    borradas = 0
    for fila in range(FILA_INICIO, ws.max_row + 1):
        for col in COLUMNAS_GEOMETRIC:
            if ws.cell(row=fila, column=col).value is not None:
                ws.cell(row=fila, column=col).value = None
                borradas += 1
    print(f"  {hoja}: {borradas} celdas limpiadas")

wb.save(EXCEL_PATH)
print("✅ Excel limpiado.")