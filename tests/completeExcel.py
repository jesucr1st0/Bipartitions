"""
 
Lee el Excel de pruebas, corre Geometric y QNodes para k=2,3,4,5
y escribe los resultados directamente en el Excel.
 
Uso:
    # Correr todo (puede tardar mucho para sistemas grandes)
    python tests/llenar_excel.py
 
    # Solo un sistema específico
    python tests/llenar_excel.py N10A
    python tests/llenar_excel.py N15B
 
    # Limitar cuántos casos por sistema (útil para pruebas rápidas)
    python tests/llenar_excel.py N10A --max 5
 
    # Solo k=2 (más rápido)
    python tests/llenar_excel.py N10A --k 2
"""
# Desactivar colorama (evita recursion depth error en Windows con unicode)
import sys as _sys
import os as _os
_sys.stdout.reconfigure(encoding='utf-8', errors='replace')
_sys.stderr.reconfigure(encoding='utf-8', errors='replace')
try:
    import colorama as _colorama
    _colorama.deinit()
except ImportError:
    pass

# Desactivar profiler antes de todo


 
import sys
import os
import time
import argparse
 
import numpy as np
import openpyxl
 
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
from models.base.application import aplicacion
aplicacion.desactivar_profiling()
 
from funcs.cargar import csv_to_tpm, excel_a_configs, excel_str_a_bits
from controllers.strategies.geometric import Geometric
from controllers.strategies.qnodes import QNodes
from controllers.strategies.kgeometric import KGeometric
from controllers.strategies.kqnodes import KQNodes
 
 
# ── Configuración ────────────────────────────────────────────────────────────
 
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
EXCEL_PATH   = os.path.join(PROJECT_ROOT, "data", "DatosPruebas2026_1.xlsx")
 
# Mapeo: nombre_sistema → nombre_hoja en el Excel
SISTEMAS = {
    "N10A": "10A-Elementos",
    "N15B": "15B-Elementos",
    "N20A": "20A-Elementos",
    "N22A": "22A-Elementos",
    "N25A": "25A-Elementos ",   # ojo: tiene espacio al final en el Excel
}
 
# Columnas donde escribir en el Excel (1-indexed, como openpyxl)
# Fila de datos empieza en la fila 6
# Estructura: [QNodes_k2, Geom_k2, QNodes_k3, Geom_k3, QNodes_k4, Geom_k4, QNodes_k5, Geom_k5]
# Cada bloque = (col_particion, col_perdida, col_tiempo)
COLUMNAS = {
    ("qnodes",    2): (4,  5,  6),
    ("geometric", 2): (7,  8,  9),
    ("qnodes",    3): (10, 11, 12),
    ("geometric", 3): (13, 14, 15),
    ("qnodes",    4): (16, 17, 18),
    ("geometric", 4): (19, 20, 21),
    ("qnodes",    5): (22, 23, 24),
    ("geometric", 5): (25, 26, 27),
}
 
FILA_INICIO_DATOS = 6   # fila 1-indexed donde empiezan los casos
 
 
# ── Helpers ──────────────────────────────────────────────────────────────────
 
def correr_estrategia(nombre, cls, tpm, k, cfg):
    """Corre una estrategia y devuelve (particion, perdida, tiempo) o None si falla."""
    try:
        obj = cls(tpm)
        if k == 2:
            r = obj.aplicar_estrategia(
                cfg["estado_inicial"],
                "1" * len(cfg["estado_inicial"]),
                cfg["alcance"],
                cfg["mecanismo"],
            )
        else:
            r = obj.find_k_mip(
                k,
                cfg["estado_inicial"],
                "1" * len(cfg["estado_inicial"]),
                cfg["alcance"],
                cfg["mecanismo"],
            )
        return str(r.particion), round(float(r.perdida), 6), round(float(r.tiempo_total), 4)
    except Exception as e:
        print(f"      [WARN] {nombre} k={k} ERROR: {e}")
        return f"ERROR: {e}", None, None
 
 
def barra_progreso(actual, total, ancho=30):
    pct = actual / total
    llenos = int(ancho * pct)
    barra = "█" * llenos + "░" * (ancho - llenos)
    return f"[{barra}] {actual}/{total} ({pct*100:.0f}%)"
 
 
# ── Función principal por sistema ─────────────────────────────────────────────
 
def procesar_sistema(sistema: str, k_vals: list[int], max_casos: int | None, wb: openpyxl.Workbook):
    if sistema not in SISTEMAS:
        print(f"  [ERROR] Sistema '{sistema}' no reconocido. Opciones: {list(SISTEMAS)}")
        return
 
    nombre_hoja = SISTEMAS[sistema]
    if nombre_hoja not in wb.sheetnames:
        print(f"  [ERROR] Hoja '{nombre_hoja}' no encontrada en el Excel.")
        return
 
    ws = wb[nombre_hoja]
 
    # Cargar TPM
    try:
        tpm = csv_to_tpm(sistema)
    except Exception as e:
        print(f"  [ERROR] No se pudo cargar la TPM de '{sistema}': {e}")
        print(f"     Asegurate de tener src/.samples/{sistema}.csv")
        return
 
    n = tpm.shape[1]
    print(f"\n{'='*60}")
    print(f"  Sistema: {sistema}  |  n={n} variables  |  hoja: {nombre_hoja}")
    print(f"{'='*60}")
 
    # Leer configuraciones del Excel
    configs = excel_a_configs(EXCEL_PATH, nombre_hoja)
    if not configs:
        print("  [ERROR] No se encontraron casos en el Excel.")
        return
 
    if max_casos:
        configs = configs[:max_casos]
 
    print(f"  Casos a procesar: {len(configs)}")
    print(f"  k valores: {k_vals}")
    print()
 
    # Estrategias por k
    estrategias = {
        2: [("qnodes", QNodes), ("geometric", Geometric)],
        3: [("qnodes", KQNodes), ("geometric", KGeometric)],
        4: [("qnodes", KQNodes), ("geometric", KGeometric)],
        5: [("qnodes", KQNodes), ("geometric", KGeometric)],
    }
 
    t_total_inicio = time.time()
 
    for i, cfg in enumerate(configs):
        fila_excel = FILA_INICIO_DATOS + i
        print(f"  Caso {i+1:>3}/{len(configs)}  |  "
              f"Alcance={cfg['alcance_excel']:<15}  Mecanismo={cfg['mecanismo_excel']}")
 
        for k in k_vals:
            if k not in estrategias:
                continue
            for nombre, cls in estrategias[k]:
                col_part, col_perd, col_tiempo = COLUMNAS[(nombre, k)]
 
                # Si ya tiene valor, saltar (para no sobreescribir trabajo previo)
                celda_actual = ws.cell(row=fila_excel, column=col_part).value
                if celda_actual is not None and str(celda_actual).startswith("ERROR") is False:
                    print(f"      [SKIP] {nombre} k={k} ya tiene resultado, saltando.")
                    continue
 
                particion, perdida, tiempo = correr_estrategia(nombre, cls, tpm, k, cfg)
 
                ws.cell(row=fila_excel, column=col_part).value  = particion
                ws.cell(row=fila_excel, column=col_perd).value  = perdida
                ws.cell(row=fila_excel, column=col_tiempo).value = tiempo
 
                estado = "[OK]" if perdida is not None else "[ERR]"
                print(f"      {estado} {nombre:<10} k={k}  "
                      f"pérdida={perdida if perdida is not None else 'ERR':>10}  "
                      f"t={tiempo if tiempo is not None else 'ERR':>8}s  "
                      f"{particion[:40] if particion else ''}")
 
        # Guardar después de cada caso para no perder progreso
        wb.save(EXCEL_PATH)
 
    t_total = time.time() - t_total_inicio
    print()
    print(f"  [OK] {sistema} completado en {t_total:.1f}s - Excel guardado.")
 
 
# ── Main ──────────────────────────────────────────────────────────────────────
 
def main():
    parser = argparse.ArgumentParser(description="Llena el Excel de pruebas con resultados.")
    parser.add_argument("sistemas", nargs="*", default=list(SISTEMAS.keys()),
                        help="Sistemas a procesar (ej: N10A N15B). Default: todos.")
    parser.add_argument("--max", type=int, default=None,
                        help="Máximo de casos por sistema (útil para pruebas rápidas).")
    parser.add_argument("--k", type=int, nargs="+", default=[2, 3, 4, 5],
                        help="Valores de k a correr. Ej: --k 2 3")
    args = parser.parse_args()
 
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] No se encontró el Excel en: {EXCEL_PATH}")
        print(f"   Cópialo a la carpeta data/ del proyecto.")
        sys.exit(1)

    print(f"[INFO] Excel: {EXCEL_PATH}")
    print(f"[INFO] Sistemas: {args.sistemas}")
    print(f"[INFO] k valores: {args.k}")
    if args.max:
        print(f"[INFO] Modo prueba: maximo {args.max} casos por sistema")
 
    wb = openpyxl.load_workbook(EXCEL_PATH)
 
    for sistema in args.sistemas:
        procesar_sistema(sistema, args.k, args.max, wb)
 
    wb.save(EXCEL_PATH)
    print()
    print("[OK] Todos los sistemas procesados. Excel final guardado.")
 
 
if __name__ == "__main__":
    main()