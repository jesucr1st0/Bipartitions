"""
Completa el Excel con Geometric/KGeometric para sistemas grandes (N20A, N22A, N25A)
donde QNodes/KQNodes no escala (n>15 -> espacio mecanismo-purview demasiado grande).

Uso:
    python tests/completar_geometric.py N20A
    python tests/completar_geometric.py N20A N22A
"""
import sys as _sys, os as _os
_sys.stdout.reconfigure(encoding='utf-8', errors='replace')
_sys.stderr.reconfigure(encoding='utf-8', errors='replace')
try:
    import colorama as _colorama
    _colorama.deinit()
except ImportError:
    pass

import sys, os, time
import numpy as np
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
from models.base.application import aplicacion
aplicacion.desactivar_profiling()

from funcs.cargar import csv_to_tpm, excel_a_configs
from controllers.strategies.geometric import Geometric
from controllers.strategies.kgeometric import KGeometric

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "data", "DatosPruebas2026_1.xlsx")

SISTEMAS = {
    "N20A": "20A-Elementos",
    "N22A": "22A-Elementos",
    "N25A": "25A-Elementos ",
}

COLUMNAS = {
    ("geometric", 2): (7, 8, 9),
    ("geometric", 3): (13, 14, 15),
    ("geometric", 4): (19, 20, 21),
    ("geometric", 5): (25, 26, 27),
}

FILA_INICIO_DATOS = 6

def correr_geometric(cls, tpm, k, cfg):
    try:
        obj = cls(tpm)
        if k == 2:
            r = obj.aplicar_estrategia(
                cfg["estado_inicial"],
                "1" * len(cfg["estado_inicial"]),
                cfg["alcance"],
                cfg["mecanismo"],
            )
        else:
            r = obj.find_k_mip(
                k, cfg["estado_inicial"],
                "1" * len(cfg["estado_inicial"]),
                cfg["alcance"], cfg["mecanismo"],
            )
        return str(r.particion), round(float(r.perdida), 6), round(float(r.tiempo_total), 4)
    except Exception as e:
        print(f"      [WARN] Geometric k={k} ERROR: {e}")
        return f"ERROR: {e}", None, None

def procesar_sistema(sistema: str, k_vals: list[int], max_casos: int | None, wb):
    if sistema not in SISTEMAS:
        print(f"  [ERROR] Sistema '{sistema}' no reconocido. Opciones: {list(SISTEMAS)}")
        return

    nombre_hoja = SISTEMAS[sistema]
    if nombre_hoja not in wb.sheetnames:
        print(f"  [ERROR] Hoja '{nombre_hoja}' no encontrada en el Excel.")
        return

    ws = wb[nombre_hoja]

    t0 = time.time()
    try:
        tpm = csv_to_tpm(sistema)
    except Exception as e:
        print(f"  [ERROR] No se pudo cargar la TPM de '{sistema}': {e}")
        return

    n = tpm.shape[1]
    print(f"\n{'='*60}")
    print(f"  Sistema: {sistema}  |  n={n} variables  |  hoja: {nombre_hoja}")
    print(f"{'='*60}")

    configs = excel_a_configs(EXCEL_PATH, nombre_hoja)
    if not configs:
        print("  [ERROR] No se encontraron casos en el Excel.")
        return

    if max_casos:
        configs = configs[:max_casos]

    print(f"  Casos a procesar: {len(configs)}")
    print(f"  k valores: {k_vals}")
    print()

    for i, cfg in enumerate(configs):
        fila_excel = FILA_INICIO_DATOS + i
        print(f"  Caso {i+1:>3}/{len(configs)}  |  "
              f"Alcance={cfg['alcance_excel']:<15}  Mecanismo={cfg['mecanismo_excel']}")

        for k in k_vals:
            col_part, col_perd, col_tiempo = COLUMNAS[("geometric", k)]
            celda_actual = ws.cell(row=fila_excel, column=col_part).value
            if celda_actual is not None and str(celda_actual).startswith("ERROR") is False:
                print(f"      [SKIP] Geometric k={k} ya tiene resultado.")
                continue

            particion, perdida, tiempo = correr_geometric(KGeometric if k > 2 else Geometric, tpm, k, cfg)

            ws.cell(row=fila_excel, column=col_part).value = particion
            ws.cell(row=fila_excel, column=col_perd).value = perdida
            ws.cell(row=fila_excel, column=col_tiempo).value = tiempo

            estado = "[OK]" if perdida is not None else "[ERR]"
            print(f"      {estado} Geometric k={k}  "
                  f"pérdida={perdida if perdida is not None else 'ERR':>10}  "
                  f"t={tiempo if tiempo is not None else 'ERR':>8}s  "
                  f"{particion[:40] if particion else ''}")

        wb.save(EXCEL_PATH)

    print(f"\n  [OK] {sistema} completado en {time.time()-t0:.1f}s")

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Completa Geometric/KGeometric para sistemas grandes.")
    parser.add_argument("sistemas", nargs="+", help="Sistemas a procesar (ej: N20A N22A)")
    parser.add_argument("--max", type=int, default=None, help="Máximo de casos (prueba rápida)")
    parser.add_argument("--k", type=int, nargs="+", default=[2, 3, 4, 5], help="Valores de k")
    args = parser.parse_args()

    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] No se encontró el Excel: {EXCEL_PATH}")
        sys.exit(1)

    print(f"[INFO] Excel: {EXCEL_PATH}")
    print(f"[INFO] Solo Geometric/KGeometric (QNodes saltado por escalamiento)")
    print(f"[INFO] Sistemas: {args.sistemas}, k={args.k}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    for sistema in args.sistemas:
        procesar_sistema(sistema, args.k, args.max, wb)
    wb.save(EXCEL_PATH)
    print("\n[OK] Todos los sistemas procesados. Excel guardado.")

if __name__ == "__main__":
    main()
